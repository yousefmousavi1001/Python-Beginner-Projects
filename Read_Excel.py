import openpyxl
class ExcelReader:
    def __init__(self,filename):
        self.filename = filename
        wb = openpyxl.load_workbook(self.filename)
        self.sheet = wb['Sheet1']
    def read_excel(self):
        data={}
        for i in self.read_first_row(): #i=kelid haye dictionari :
            data[i]=[]
        for row in range(2,self.sheet.max_row+1):
            for column, i  in enumerate(self.read_first_row(),start=1):
                value=self.sheet.cell(row=row,column=column).value
                data[i].append(value)
        return data
    def read_first_row(self):
        f_r=[]
        for i in range(1,self.sheet.max_column+1):
            f_r.append(self.sheet.cell(row=1,column=i).value)
        return f_r
    def read_last_column(self):
        l_c=[]
        for i in range(1,self.sheet.max_row+1):
            l_c.append(self.sheet.cell(row=i,column=self.sheet.max_column).value)
        return l_c
    def read_row(self,row):
        l_k = []
        for i in range(1, self.sheet.max_column ):
            l_k.append(self.sheet.cell(row=row, column=i).value)
        return l_k
    def read_column(self,column):
        l_s = []
        for i in range(1, self.sheet.max_row ):
            l_s.append(self.sheet.cell(row=column, column=i).value)
        return l_s
